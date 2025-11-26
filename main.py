import asyncio
import base64
import glob
import io
import os
import threading
import time
import uuid
from contextlib import asynccontextmanager
from datetime import datetime

import cv2
import numpy as np
from fastapi import FastAPI, File, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from PIL import Image
import supervision as sv
import uvicorn

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl не установлен. Будет использоваться текстовый формат.")

from core.image_processor import (
    add_labels_to_image,
    create_combined_mask_image,
    create_detections_from_results,
    create_overlay_image,
    draw_masks_and_remember_colors,
    get_labels_from_predictions,
    get_polygon_colors,
    load_and_convert_image,
    save_image_to_tmp,
    save_multiple_formats
)
from core.model_manager import (
    apply_class_mapping,
    apply_damage_class_mapping,
    create_combined_result,
    create_damage_combined_result,
    create_full_damage_combined_result,
    filter_predictions,
    get_damage_predictions_async,
    get_full_union_predictions_async,
    get_predictions_async,
    initialize_damage_models,
    initialize_models
)

# Глобальное хранилище результатов обработки по file_id
processed_results = {}

# Списки допустимых классов из model_manager.py
PARTS_CLASSES = {
    # Основная модель
    "КРЫЛО П", "БОКОВИНА В СБ З", "СТЕКЛО ЛОБОВОЕ", "СТЕКЛО ЗАДНЕЕ",
    "НОМЕРНОЙ ЗНАК З", "НОМЕРНОЙ ЗНАК П", "БАМПЕР", "КАПОТ", "РАДИАТОР",
    # Модель колес
    "КОЛЕСНЫЙ ДИСК", "ШИНА", "КРЫША", "Стекло", "ЭМБЛЕМА ПРОИЗВОДИТЕЛЯ", "ЛЮЧОК Т/БАКА",
    # Модель дверей
    "ДВЕРЬ З Л", "ДВЕРЬ З ПР", "ГАБ ФОНАРЬ З ПР", "ГАБ ФОНАРЬ З Л",
    "ДВЕРЬ П Л", "ДВЕР П ПР", "ФАРА В СБОРЕ Л", "ФАРА В СБОРЕ ПР",
    "ЗЕРКАЛО НАР Л", "ЗЕРКАЛО НАР ПР"
}

DAMAGE_CLASSES = {
    "Складка", "Утрата фрагментов", "Трещина", "Царапина", "Вмятина", "Поверхностная коррозия"
}


def clear_old_processed_results():
    """Очищает старые результаты обработки (старше 1 часа)."""
    current_time = time.time()
    max_age_hours = 1  # Удаляем результаты старше 1 часа

    to_remove = []
    for file_id, result in processed_results.items():
        try:
            timestamp = result.get('timestamp', '')
            if timestamp:
                # Преобразуем ISO timestamp в timestamp
                from datetime import datetime
                result_time = datetime.fromisoformat(timestamp).timestamp()
                if current_time - result_time > max_age_hours * 3600:
                    to_remove.append(file_id)
        except:
            # Если не можем распарсить timestamp, считаем устаревшим
            to_remove.append(file_id)

    for file_id in to_remove:
        del processed_results[file_id]
        print(f"🗑️ Удален старый результат: {file_id}")

    if to_remove:
        print(f"✅ Очищено {len(to_remove)} старых результатов")


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Управление жизненным циклом приложения."""
    # Startup
    print("🚀 Запуск DCD Vision Web App...")
    print("🧹 Очистка старых временных файлов...")
    cleanup_temp_files()
    print("🗑️ Очистка старых результатов обработки...")
    clear_old_processed_results()
    print("🔄 Инициализация моделей в фоне...")
    preload_models()

    yield

    # Shutdown
    print("🛑 Завершение работы DCD Vision Web App...")
    try:
        from core.model_manager import save_models_to_cache
        save_models_to_cache()
        print("💾 Кэш моделей сохранен при завершении")
    except Exception as e:
        print(f"⚠️ Ошибка при сохранении кэша: {e}")


def cleanup_temp_files():
    """Очистка старых временных файлов."""
    try:
        # Ищем все временные файлы в корне проекта
        temp_files = glob.glob("temp_*.jpg") + glob.glob("temp_damage_*.jpg")

        current_time = time.time()
        max_age_hours = 1  # Удаляем файлы старше 1 часа

        cleaned_count = 0
        for temp_file in temp_files:
            if os.path.isfile(temp_file):
                file_age_hours = (current_time - os.path.getmtime(temp_file)) / 3600
                if file_age_hours > max_age_hours:
                    try:
                        os.remove(temp_file)
                        cleaned_count += 1
                        print(f"🗑️ Удален старый временный файл: {temp_file}")
                    except Exception as e:
                        print(f"⚠️ Не удалось удалить {temp_file}: {e}")

        if cleaned_count > 0:
            print(f"✅ Очищено {cleaned_count} старых временных файлов")

        # Также проверяем и очищаем файлы в папке tmp
        tmp_files = glob.glob("tmp/temp_*.jpg") + glob.glob("tmp/temp_damage_*.jpg")

        for tmp_file in tmp_files:
            if os.path.isfile(tmp_file):
                file_age_hours = (current_time - os.path.getmtime(tmp_file)) / 3600
                if file_age_hours > max_age_hours:
                    try:
                        os.remove(tmp_file)
                        print(f"🗑️ Удален старый файл из tmp: {tmp_file}")
                    except Exception as e:
                        print(f"⚠️ Не удалось удалить {tmp_file}: {e}")

    except Exception as e:
        print(f"⚠️ Ошибка при очистке временных файлов: {e}")


def preload_models():
    """Предварительная загрузка всех моделей в отдельном потоке."""
    def _preload():
        try:
            print("🔄 Начинаем предварительную загрузку моделей...")
            # Загружаем модели деталей автомобиля
            initialize_models()
            # Загружаем модели повреждений
            initialize_damage_models()
            print("✅ Все модели успешно загружены и закэшированы")
        except Exception as e:
            print(f"⚠️ Ошибка при предварительной загрузке моделей: {e}")
            print("Модели будут загружены при первом использовании")

    # Запускаем в отдельном потоке, чтобы не блокировать запуск сервера
    preload_thread = threading.Thread(target=_preload, daemon=True)
    preload_thread.start()


app = FastAPI(title="DCD Vision Web App", lifespan=lifespan)

# Создаем директорию tmp если она не существует
os.makedirs("tmp", exist_ok=True)
print(f"📁 Директория tmp: {'✅ существует' if os.path.exists('tmp') else '❌ не существует'}")
print(f"📝 Директория tmp: {'✅ доступна для записи' if os.access('tmp', os.W_OK) else '❌ недоступна для записи'}")

# Монтируем статические файлы с абсолютными путями
app.mount("/static", StaticFiles(directory="static"), name="static")
app.mount("/tmp", StaticFiles(directory="tmp"), name="tmp")

# Фавикон
@app.get("/favicon.ico")
async def favicon():
    """Возвращает favicon."""
    from fastapi.responses import Response
    # Возвращаем пустой ответ с правильным статусом чтобы убрать ошибку 404
    return Response(content="", status_code=204)

# Настраиваем шаблоны
templates = Jinja2Templates(directory="templates")


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    """Главная страница с drag'n'drop интерфейсом."""
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/upload")
async def upload_image(file: UploadFile = File(...)):
    """Обработка загруженного изображения."""
    try:
        # Читаем файл
        contents = await file.read()
        image = Image.open(io.BytesIO(contents))

        # Генерируем уникальный ID для файла
        file_id = str(uuid.uuid4())[:8]

        # Сохраняем временно для обработки в папке tmp
        temp_path = os.path.join("tmp", f"temp_{file_id}_{int(time.time())}.jpg")
        image.save(temp_path, "JPEG")

        # Обрабатываем изображение
        print(f"🔄 Начинаем обработку файла: {temp_path}")
        result_info = await process_car_image_async(temp_path, file_id, file.filename)
        print(f"✅ Обработка завершена, результат: {result_info}")

        # Удаляем временный файл
        if os.path.exists(temp_path):
            os.remove(temp_path)
            print(f"🗑️ Удален временный файл: {temp_path}")
        else:
            print(f"⚠️ Временный файл не найден для удаления: {temp_path}")

        # Преобразуем пути в URL для фронтенда
        overlay_url = f"/tmp/{os.path.basename(result_info['overlay'])}"
        mask_url = f"/tmp/{os.path.basename(result_info['mask'])}"
        original_url = f"/tmp/{os.path.basename(result_info['original'])}"

        # Проверяем, что файлы действительно существуют
        overlay_path = result_info['overlay']
        mask_path = result_info['mask']
        original_path = result_info['original']

        print(f"📁 Проверяем файлы:")
        print(f"   Overlay: {overlay_path} - {'✅' if os.path.exists(overlay_path) else '❌'}")
        print(f"   Mask: {mask_path} - {'✅' if os.path.exists(mask_path) else '❌'}")
        print(f"   Original: {original_path} - {'✅' if os.path.exists(original_path) else '❌'}")

        print(f"📤 Отправляем на фронтенд (детали):")
        print(f"   File ID: {result_info.get('file_id', file_id)}")
        print(f"   Polygons count: {len(result_info.get('polygons', []))}")
        print(f"   Detections count: {len(result_info.get('detections', []))}")

        return {
            "success": True,
            "overlay": overlay_url,
            "mask": mask_url,
            "original": original_url,
            "filename": file.filename,
            "file_id": file_id,
            "detections": result_info.get('detections', []),
            "polygons": result_info.get('polygons', []),
            "models_info": result_info.get('models_info', {})
        }

    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


@app.post("/upload_damage")
async def upload_damage_image(file: UploadFile = File(...)):
    """Обработка загруженного изображения для сегментации повреждений."""
    try:
        # Читаем файл
        contents = await file.read()
        image = Image.open(io.BytesIO(contents))

        # Генерируем уникальный ID для файла
        file_id = str(uuid.uuid4())[:8]

        # Сохраняем временно для обработки в папке tmp
        temp_path = os.path.join("tmp", f"temp_damage_{file_id}_{int(time.time())}.jpg")
        image.save(temp_path, "JPEG")

        # Обрабатываем изображение повреждений
        result_info = await process_damage_image_async(temp_path, file_id, file.filename)

        # Удаляем временный файл
        os.remove(temp_path)

        # Преобразуем пути в URL для фронтенда
        overlay_url = f"/tmp/{os.path.basename(result_info['overlay'])}"
        mask_url = f"/tmp/{os.path.basename(result_info['mask'])}"
        original_url = f"/tmp/{os.path.basename(result_info['original'])}"

        print(f"📤 Отправляем на фронтенд (повреждения):")
        print(f"   File ID: {result_info.get('file_id', file_id)}")
        print(f"   Polygons count: {len(result_info.get('polygons', []))}")
        print(f"   Detections count: {len(result_info.get('detections', []))}")

        return {
            "success": True,
            "overlay": overlay_url,
            "mask": mask_url,
            "original": original_url,
            "filename": file.filename,
            "file_id": file_id,
            "detections": result_info.get('detections', []),
            "polygons": result_info.get('polygons', []),
            "models_info": result_info.get('models_info', {})
        }

    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


async def process_car_image_async(path, file_id=None, original_filename=None):
    """
    Полностью асинхронный пайплайн обработки изображения автомобиля.

    Args:
        path (str): Путь к изображению
        file_id (str): Уникальный ID файла
        original_filename (str): Оригинальное имя файла, загруженное пользователем

    Returns:
        dict: Информация о сохраненных файлах (overlay, mask, filename)
    """
    start_total_time = time.time()

    print(f"🚗 Начинаем обработку: {path}")

    # Одновременные асинхронные запросы ко всем моделям
    wheels_result, doors_result, main_result = await get_predictions_async(path)

    # Фильтрация предсказаний
    filtered_main, filtered_wheels, filtered_doors = filter_predictions(
        wheels_result, doors_result, main_result)

    # Переименование классов
    filtered_main, filtered_wheels, filtered_doors = apply_class_mapping(
        filtered_main, filtered_wheels, filtered_doors)

    # Создание объединенного результата
    combined_result = create_combined_result(
        filtered_main, filtered_wheels, filtered_doors, main_result)

    # Собираем все уникальные классы из ОТФИЛЬТРОВАННЫХ результатов
    all_filtered_classes = set()

    # Из отфильтрованной wheels модели
    for pred in filtered_wheels:
        if 'class' in pred:
            all_filtered_classes.add(pred['class'])

    # Из отфильтрованной doors модели
    for pred in filtered_doors:
        if 'class' in pred:
            all_filtered_classes.add(pred['class'])

    # Из отфильтрованной main модели
    for pred in filtered_main:
        if 'class' in pred:
            all_filtered_classes.add(pred['class'])

    # Конвертируем в список и сортируем
    detected_classes = sorted(list(all_filtered_classes))

    print(f"🎯 Классы ПОСЛЕ ФИЛЬТРАЦИИ (только те, что попали на изображение): {detected_classes}")
    print(f"📊 Всего отфильтрованных уникальных классов: {len(detected_classes)}")
    print(f"📤 Отправляем в frontend: models_info.detected_classes = {detected_classes}")

    # Проверяем, что есть предсказания перед созданием детекций
    if not combined_result.get('predictions'):
        print("⚠️ Нет предсказаний от моделей деталей, создаем пустые детекции")
        # Создаем пустые детекции если нет предсказаний
        detections = sv.Detections.empty()
    else:
        # Создание детекций
        detections = create_detections_from_results(combined_result)

    print(f"🎨 Детекции с цветами: {len(detections)} шт.")

    # Сохраняем информацию о всех моделях для frontend
    models_info = {
        'detected_classes': detected_classes,
        'predictions': combined_result.get('predictions', []),  # Добавляем предсказания для создания отдельных масок
        'models_stats': {
            'wheels_model': {
                'name': 'Колеса и шины',
                'predictions_count': len(wheels_result.get('predictions', [])),
                'filtered_count': len(filtered_wheels)
            },
            'doors_model': {
                'name': 'Двери и зеркала',
                'predictions_count': len(doors_result.get('predictions', [])),
                'filtered_count': len(filtered_doors)
            },
            'main_model': {
                'name': 'Основные детали',
                'predictions_count': len(main_result.get('predictions', [])),
                'filtered_count': len(filtered_main)
            }
        }
    }

    # Получение лейблов
    labels = get_labels_from_predictions(combined_result["predictions"])

    # Загрузка изображения
    original_image = load_and_convert_image(path)

    # Словарь для хранения цветов полигонов
    polygon_colors = get_polygon_colors()

    # Создаем изображение с наложенными масками
    annotated_image = draw_masks_and_remember_colors(
        original_image.copy(), detections, labels, polygon_colors)

    # Ручное добавление лейблов через PIL
    pil_image = Image.fromarray(annotated_image)
    pil_image = add_labels_to_image(pil_image, detections, labels, polygon_colors)

    # Создаем чистое изображение масок
    mask_image = create_combined_mask_image(detections, labels, original_image.shape, polygon_colors)

    end_total_time = time.time()
    total_time = end_total_time - start_total_time

    print(".2f")
    print(f"📊 Обработано {len(labels)} объектов на изображении")

    # Сохраняем результат в нескольких форматах
    filename_to_save = original_filename or os.path.basename(path)
    result_info = save_multiple_formats(
        np.array(pil_image),
        mask_image,
        file_id,
        filename_to_save,
        detections,
        labels,
        models_info,
        polygon_colors,
        original_image
    )

    print(f"💾 Результаты сохранены: {result_info}")

    # Добавляем file_id для работы с масками в frontend
    result_info['file_id'] = file_id

    # Сохраняем результат для экспорта - все из модели деталей идут как parts
    polygons_for_export = []
    for i, detection in enumerate(result_info.get('detections', [])):
        polygons_for_export.append({
            'id': i + 1,
            'class': detection.get('class', ''),
            'confidence': detection.get('confidence', 0.0),
            'bbox': detection.get('bbox', [0, 0, 0, 0]),
            'source': 'parts'  # Все детекции из модели деталей
        })

    processed_results[file_id] = {
        'filename': filename_to_save,
        'mode': 'parts',
        'polygons': polygons_for_export,
        'models_info': models_info,
        'timestamp': datetime.now().isoformat()
    }

    return result_info


async def process_damage_image_async(path, file_id=None, original_filename=None):
    """
    Полностью асинхронный пайплайн обработки повреждений автомобиля.

    Args:
        path (str): Путь к изображению
        file_id (str): Уникальный ID файла
        original_filename (str): Оригинальное имя файла, загруженное пользователем

    Returns:
        dict: Информация о сохраненных файлах (overlay, mask, filename)
    """
    start_total_time = time.time()

    print(f"🚨 Начинаем обработку повреждений: {path}")

    # Одновременные асинхронные запросы ко всем моделям повреждений
    damage_result_1, damage_result_2, damage_result_3, damage_result_4 = await get_damage_predictions_async(path)

    # Применяем маппинг классов повреждений и фильтрацию
    damage_result_1, damage_result_2, damage_result_3, damage_result_4 = apply_damage_class_mapping(
        damage_result_1, damage_result_2, damage_result_3, damage_result_4)

    # Создание объединенного результата повреждений
    combined_result = create_damage_combined_result(
        damage_result_1, damage_result_2, damage_result_3, damage_result_4)

    # Собираем все уникальные классы из объединенного результата повреждений
    all_damage_classes = set()

    # Из объединенного результата (уже содержит все предсказания от всех моделей)
    for pred in combined_result.get('predictions', []):
        if 'class' in pred:
            all_damage_classes.add(pred['class'])

    # Конвертируем в список и сортируем
    detected_damage_classes = sorted(list(all_damage_classes))

    print(f"🎯 Классы повреждений (все найденные от всех моделей): {detected_damage_classes}")
    print(f"📊 Всего уникальных классов повреждений: {len(detected_damage_classes)}")
    print(f"📤 Отправляем в frontend: models_info.detected_classes = {detected_damage_classes}")

    # Проверяем, что есть предсказания перед созданием детекций
    if not combined_result.get('predictions'):
        print("⚠️ Нет предсказаний от моделей повреждений, создаем пустые детекции")
        # Создаем пустые детекции если нет предсказаний
        detections = sv.Detections.empty()
    else:
        # Создание детекций
        detections = create_detections_from_results(combined_result)

    print(f"🎨 Детекции повреждений с цветами: {len(detections)} шт.")

    # Сохраняем информацию о всех моделях повреждений для frontend
    models_info = {
        'detected_classes': detected_damage_classes,
        'predictions': combined_result.get('predictions', []),  # Добавляем предсказания для создания отдельных масок
        'models_stats': {
            'damage_model_1': {
                'name': 'Обнаружение повреждений (FRMNL)',
                'predictions_count': len(damage_result_1.get('predictions', [])),
                'filtered_count': len(damage_result_1.get('predictions', []))
            },
            'damage_model_2': {
                'name': 'Обнаружение повреждений (VYHVW)',
                'predictions_count': len(damage_result_2.get('predictions', [])),
                'filtered_count': len(damage_result_2.get('predictions', []))
            },
            'damage_model_3': {
                'name': 'Обнаружение повреждений (Bilgi)',
                'predictions_count': len(damage_result_3.get('predictions', [])),
                'filtered_count': len(damage_result_3.get('predictions', []))
            }
        }
    }

    # Получение лейблов
    labels = get_labels_from_predictions(combined_result["predictions"])

    # Загрузка изображения
    original_image = load_and_convert_image(path)

    # Словарь для хранения цветов полигонов
    polygon_colors = get_polygon_colors()

    # Создаем изображение с наложенными масками
    annotated_image = draw_masks_and_remember_colors(
        original_image.copy(), detections, labels, polygon_colors)

    # Ручное добавление лейблов через PIL
    pil_image = Image.fromarray(annotated_image)
    pil_image = add_labels_to_image(pil_image, detections, labels, polygon_colors)

    # Создаем чистое изображение масок
    mask_image = create_combined_mask_image(detections, labels, original_image.shape, polygon_colors)

    end_total_time = time.time()
    total_time = end_total_time - start_total_time

    print(".2f")
    print(f"📊 Обработано {len(labels)} повреждений на изображении")

    # Сохраняем результат в нескольких форматах
    filename_to_save = original_filename or os.path.basename(path)
    result_info = save_multiple_formats(
        np.array(pil_image),
        mask_image,
        file_id,
        filename_to_save,
        detections,
        labels,
        models_info,
        polygon_colors,
        original_image
    )

    print(f"💾 Результаты повреждений сохранены: {result_info}")

    # Добавляем file_id для работы с масками в frontend
    result_info['file_id'] = file_id

    # Сохраняем результат для экспорта - все из модели повреждений идут как damage
    polygons_for_export = []
    for i, detection in enumerate(result_info.get('detections', [])):
        polygons_for_export.append({
            'id': i + 1,
            'class': detection.get('class', ''),
            'confidence': detection.get('confidence', 0.0),
            'bbox': detection.get('bbox', [0, 0, 0, 0]),
            'source': 'damage'  # Все детекции из модели повреждений
        })

    processed_results[file_id] = {
        'filename': filename_to_save,
        'mode': 'damage',
        'polygons': polygons_for_export,
        'models_info': models_info,
        'timestamp': datetime.now().isoformat()
    }

    return result_info


@app.post("/upload_full_union")
async def upload_full_union_image(file: UploadFile = File(...)):
    """Обработка загруженного изображения для полного объединения всех моделей (детали + повреждения)."""
    try:
        # Читаем файл
        contents = await file.read()
        image = Image.open(io.BytesIO(contents))

        # Генерируем уникальный ID для файла
        file_id = str(uuid.uuid4())[:8]

        # Сохраняем временно для обработки в папке tmp
        temp_path = os.path.join("tmp", f"temp_full_union_{file_id}_{int(time.time())}.jpg")
        image.save(temp_path, "JPEG")

        # Обрабатываем изображение с полным объединением
        result_info = await process_full_union_image_async(temp_path, file_id, file.filename)
        print(f"✅ Обработка полного объединения завершена, результат: {result_info}")

        # Удаляем временный файл
        if os.path.exists(temp_path):
            os.remove(temp_path)
            print(f"🗑️ Удален временный файл: {temp_path}")

        # Преобразуем пути в URL для фронтенда
        overlay_url = f"/tmp/{os.path.basename(result_info['overlay'])}"
        mask_url = f"/tmp/{os.path.basename(result_info['mask'])}"
        original_url = f"/tmp/{os.path.basename(result_info['original'])}"

        # Проверяем, что файлы действительно существуют
        overlay_path = result_info['overlay']
        mask_path = result_info['mask']
        original_path = result_info['original']

        print(f"📁 Проверяем файлы полного объединения:")
        print(f"   Overlay: {overlay_path} - {'✅' if os.path.exists(overlay_path) else '❌'}")
        print(f"   Mask: {mask_path} - {'✅' if os.path.exists(mask_path) else '❌'}")
        print(f"   Original: {original_path} - {'✅' if os.path.exists(original_path) else '❌'}")

        print(f"📤 Отправляем на фронтенд (полное объединение):")
        print(f"   File ID: {result_info.get('file_id', file_id)}")
        print(f"   Polygons count: {len(result_info.get('polygons', []))}")
        print(f"   Detections count: {len(result_info.get('detections', []))}")

        return {
            "success": True,
            "overlay": overlay_url,
            "mask": mask_url,
            "original": original_url,
            "filename": file.filename,
            "file_id": file_id,
            "detections": result_info.get('detections', []),
            "polygons": result_info.get('polygons', []),
            "models_info": result_info.get('models_info', {})
        }

    except Exception as e:
        print(f"❌ Ошибка при обработке полного объединения: {str(e)}")
        import traceback
        traceback.print_exc()

        return {
            "success": False,
            "error": f"Ошибка обработки: {str(e)}"
        }


async def process_full_union_image_async(path, file_id=None, original_filename=None):
    """
    Полностью асинхронный пайплайн обработки полного объединения автомобиля.

    Args:
        path (str): Путь к изображению
        file_id (str): Уникальный ID файла
        original_filename (str): Оригинальное имя файла, загруженное пользователем

    Returns:
        dict: Информация о сохраненных файлах (overlay, mask, filename)
    """
    start_total_time = time.time()

    print(f"🚗 Начинаем обработку полного объединения (детали + повреждения): {path}")

    # Одновременные асинхронные запросы ко всем моделям (детали + повреждения)
    wheels_result, doors_result, main_result, damage_result_1, damage_result_2, damage_result_3, damage_result_4 = await get_full_union_predictions_async(path)

    # Фильтрация предсказаний по классам
    filtered_main, filtered_wheels, filtered_doors = filter_predictions(wheels_result, doors_result, main_result)

    # Применяем переименование классов для деталей
    filtered_main, filtered_wheels, filtered_doors = apply_class_mapping(filtered_main, filtered_wheels, filtered_doors)

    # Применяем маппинг классов повреждений и фильтрацию
    damage_result_1, damage_result_2, damage_result_3, damage_result_4 = apply_damage_class_mapping(
        damage_result_1, damage_result_2, damage_result_3, damage_result_4)

    # Получаем отфильтрованные результаты повреждений
    filtered_damage_1 = damage_result_1.get("predictions", [])
    filtered_damage_2 = damage_result_2.get("predictions", [])
    filtered_damage_3 = damage_result_3.get("predictions", [])
    filtered_damage_4 = damage_result_4.get("predictions", [])

    # Создание объединенного результата для полного объединения
    combined_result = create_full_damage_combined_result(
        {"predictions": filtered_main, "image": main_result.get("image", {})},
        {"predictions": filtered_wheels, "image": wheels_result.get("image", {})},
        {"predictions": filtered_doors, "image": doors_result.get("image", {})},
        {"predictions": filtered_damage_1, "image": damage_result_1.get("image", {})},
        {"predictions": filtered_damage_2, "image": damage_result_2.get("image", {})},
        {"predictions": filtered_damage_3, "image": damage_result_3.get("image", {})},
        {"predictions": filtered_damage_4, "image": damage_result_4.get("image", {})})

    # Собираем все уникальные классы из объединенного результата (детали + повреждения)
    all_classes = set()

    # Из объединенного результата (уже содержит все предсказания от всех моделей)
    for pred in combined_result.get('predictions', []):
        if 'class' in pred:
            all_classes.add(pred['class'])

    # Конвертируем в список и сортируем
    detected_classes = sorted(list(all_classes))

    print(f"🎯 Классы полного объединения (детали + повреждения): {detected_classes}")
    print(f"📊 Всего уникальных классов: {len(detected_classes)}")
    print(f"📤 Отправляем в frontend: models_info.detected_classes = {detected_classes}")

    # Проверяем, что есть предсказания перед созданием детекций
    if not combined_result.get('predictions'):
        print("⚠️ Нет предсказаний от моделей, создаем пустые детекции")
        # Создаем пустые детекции если нет предсказаний
        detections = sv.Detections.empty()
    else:
        # Создание детекций
        detections = create_detections_from_results(combined_result)

    print(f"🎨 Детекции с цветами: {len(detections)} шт.")

    # Сохраняем информацию о всех моделях для frontend
    models_info = {
        'detected_classes': detected_classes,
        'predictions': combined_result.get('predictions', []),  # Добавляем предсказания для создания отдельных масок
        'models_stats': {
            'wheels_model': {
                'name': 'Колеса и шины',
                'predictions_count': len(wheels_result.get('predictions', [])),
                'filtered_count': len(filtered_wheels)
            },
            'doors_model': {
                'name': 'Двери и зеркала',
                'predictions_count': len(doors_result.get('predictions', [])),
                'filtered_count': len(filtered_doors)
            },
            'main_model': {
                'name': 'Основные детали',
                'predictions_count': len(main_result.get('predictions', [])),
                'filtered_count': len(filtered_main)
            },
            'damage_model_1': {
                'name': 'Обнаружение повреждений (FRMNL)',
                'predictions_count': len(damage_result_1.get('predictions', [])),
                'filtered_count': len(filtered_damage_1)
            },
            'damage_model_2': {
                'name': 'Обнаружение повреждений (VYHVW)',
                'predictions_count': len(damage_result_2.get('predictions', [])),
                'filtered_count': len(filtered_damage_2)
            },
            'damage_model_3': {
                'name': 'Обнаружение повреждений (Bilgi)',
                'predictions_count': len(damage_result_3.get('predictions', [])),
                'filtered_count': len(filtered_damage_3)
            },
            'damage_model_4': {
                'name': 'Коррозия и ржавчина',
                'predictions_count': len(damage_result_4.get('predictions', [])),
                'filtered_count': len(filtered_damage_4)
            }
        }
    }

    # Получение лейблов
    labels = get_labels_from_predictions(combined_result["predictions"])

    # Загрузка изображения
    original_image = load_and_convert_image(path)

    # Словарь для хранения цветов полигонов
    polygon_colors = get_polygon_colors()

    # Создаем изображение с наложенными масками
    annotated_image = draw_masks_and_remember_colors(
        original_image.copy(), detections, labels, polygon_colors)

    # Ручное добавление лейблов через PIL
    pil_image = Image.fromarray(annotated_image)
    pil_image = add_labels_to_image(pil_image, detections, labels, polygon_colors)

    # Создаем чистое изображение масок
    mask_image = create_combined_mask_image(detections, labels, original_image.shape, polygon_colors)

    end_total_time = time.time()
    total_time = end_total_time - start_total_time

    print(".2f")
    print(f"📊 Обработано {len(labels)} объектов (детали + повреждения) на изображении")

    # Сохраняем результат в нескольких форматах
    filename_to_save = original_filename or os.path.basename(path)
    result_info = save_multiple_formats(
        np.array(pil_image),
        mask_image,
        file_id,
        filename_to_save,
        detections,
        labels,
        models_info,
        polygon_colors,
        original_image
    )

    print(f"💾 Результаты полного объединения сохранены: {result_info}")

    # Добавляем file_id для работы с масками в frontend
    result_info['file_id'] = file_id

    # Сохраняем результат для экспорта - определяем по спискам классов
    polygons_for_export = []
    for i, detection in enumerate(result_info.get('detections', [])):
        class_name = detection.get('class', '')
        source = 'damage' if class_name in DAMAGE_CLASSES else 'parts'

        polygons_for_export.append({
            'id': i + 1,
            'class': class_name,
            'confidence': detection.get('confidence', 0.0),
            'bbox': detection.get('bbox', [0, 0, 0, 0]),
            'source': source
        })

    processed_results[file_id] = {
        'filename': filename_to_save,
        'mode': 'full-union',
        'polygons': polygons_for_export,
        'models_info': models_info,
        'timestamp': datetime.now().isoformat()
    }

    return result_info


def create_full_union_result(parts_result, damage_result):
    """
    Объединяет результаты деталей и повреждений в один результат.

    Args:
        parts_result (dict): Результат обработки деталей
        damage_result (dict): Результат обработки повреждений

    Returns:
        dict: Объединенный результат
    """
    print("🔗 Создаем полное объединение результатов...")

    # Объединяем все предсказания
    all_predictions = []

    # Добавляем предсказания деталей
    if 'predictions' in parts_result:
        for pred in parts_result['predictions']:
            pred_copy = pred.copy()
            pred_copy['source'] = 'parts'  # Помечаем источник
            all_predictions.append(pred_copy)

    # Добавляем предсказания повреждений
    if 'predictions' in damage_result:
        for pred in damage_result['predictions']:
            pred_copy = pred.copy()
            pred_copy['source'] = 'damage'  # Помечаем источник
            all_predictions.append(pred_copy)

    # Создаем объединенный результат
    union_result = {
        'predictions': all_predictions,
        'image': parts_result.get('image', damage_result.get('image', {})),
        'success': True
    }

    print(f"✅ Полное объединение создано: {len(all_predictions)} предсказаний")

    # Отладка: проверяем классы в объединенных предсказаниях
    parts_classes = set(p.get('class', '') for p in all_predictions if p.get('source') == 'parts')
    damage_classes = set(p.get('class', '') for p in all_predictions if p.get('source') == 'damage')
    print(f"📊 Классы деталей: {sorted(parts_classes)}")
    print(f"📊 Классы повреждений: {sorted(damage_classes)}")

    return union_result





@app.post("/composite_image")
async def composite_image(request: Request):
    """
    Создание композитного изображения из видимых полигонов.

    Args:
        request: Запрос с данными о видимых полигонах

    Returns:
        dict: Путь к композитному изображению
    """
    try:
        data = await request.json()
        file_id = data.get('file_id')
        visible_polygon_ids = data.get('visible_polygons', [])
        use_colored = data.get('use_colored', True)

        if not file_id or not visible_polygon_ids:
            return {"success": False, "error": "Missing file_id or visible_polygons"}

        # Найти директорию с масками полигонов
        masks_dir = None
        for item in os.listdir("tmp"):
            if item.startswith(f"masks_{file_id}_"):
                masks_dir = os.path.join("tmp", item)
                break

        if not masks_dir or not os.path.exists(masks_dir):
            return {"success": False, "error": "Masks directory not found"}

        # Загрузить оригинальное изображение
        original_path = os.path.join("tmp", f"original_{file_id}.jpg")
        if not os.path.exists(original_path):
            return {"success": False, "error": "Original image not found"}

        original_image = load_and_convert_image(original_path)

        # Создать композитное изображение
        composite_image = np.copy(original_image)

        # Наложить видимые полигоны
        for polygon_id in visible_polygon_ids:
            if use_colored:
                mask_path = os.path.join(masks_dir, f"{polygon_id}_colored.png")
            else:
                mask_path = os.path.join(masks_dir, f"{polygon_id}_binary.png")

            if os.path.exists(mask_path):
                mask_image = cv2.imread(mask_path, cv2.IMREAD_UNCHANGED)
                if mask_image is not None:
                    # Наложить маску на композитное изображение
                    if use_colored:
                        # Для цветных масок - применяем полупрозрачность
                        # Сначала проверяем, есть ли альфа-канал
                        if mask_image.shape[2] == 4:
                            # Маска с альфа-каналом - конвертируем в RGB
                            mask_rgb = cv2.cvtColor(mask_image, cv2.COLOR_BGRA2BGR)
                        else:
                            mask_rgb = mask_image

                        # Накладываем маску с прозрачностью
                        composite_image = cv2.addWeighted(composite_image, 1.0, mask_rgb, 0.7, 0)
                    else:
                        # Для бинарных масок - применить как маску
                        if mask_image.shape[2] == 4:
                            mask_binary = mask_image[:, :, 3] > 128  # Используем альфа-канал
                        else:
                            mask_binary = cv2.cvtColor(mask_image, cv2.COLOR_BGR2GRAY) > 128

                        # Создаем полупрозрачную маску
                        overlay = composite_image.copy()
                        overlay[mask_binary] = [255, 0, 0]  # Красный цвет для видимых полигонов
                        composite_image = cv2.addWeighted(composite_image, 0.7, overlay, 0.3, 0)

        # Сохранить композитное изображение
        timestamp = int(time.time())
        composite_filename = f"composite_{file_id}_{timestamp}.jpg"
        composite_path = os.path.join("tmp", composite_filename)

        pil_composite = Image.fromarray(cv2.cvtColor(composite_image, cv2.COLOR_BGR2RGB))
        pil_composite.save(composite_path)

        return {
            "success": True,
            "composite_image": composite_path
        }

    except Exception as e:
        print(f"Error creating composite image: {e}")
        return {"success": False, "error": str(e)}


@app.post("/composite_mask")
async def composite_mask(request: Request):
    """
    Создание композитной маски из видимых полигонов.

    Args:
        request: Запрос с данными о видимых полигонах

    Returns:
        dict: Путь к композитной маске
    """
    try:
        data = await request.json()
        file_id = data.get('file_id')
        visible_polygon_ids = data.get('visible_polygons', [])

        if not file_id or not visible_polygon_ids:
            return {"success": False, "error": "Missing file_id or visible_polygons"}

        # Найти директорию с масками полигонов
        masks_dir = None
        for item in os.listdir("tmp"):
            if item.startswith(f"masks_{file_id}_"):
                masks_dir = os.path.join("tmp", item)
                break

        if not masks_dir or not os.path.exists(masks_dir):
            return {"success": False, "error": "Masks directory not found"}

        # Получить размер изображения из любой маски
        mask_files = [f for f in os.listdir(masks_dir) if f.endswith('_binary.png')]
        if not mask_files:
            return {"success": False, "error": "No mask files found"}

        sample_mask_path = os.path.join(masks_dir, mask_files[0])
        sample_mask = cv2.imread(sample_mask_path, cv2.IMREAD_GRAYSCALE)
        if sample_mask is None:
            return {"success": False, "error": "Cannot read mask file"}

        height, width = sample_mask.shape[:2]

        # Создать композитную маску
        composite_mask = np.zeros((height, width, 3), dtype=np.uint8)

        # Наложить видимые полигоны
        for polygon_id in visible_polygon_ids:
            colored_mask_path = os.path.join(masks_dir, f"{polygon_id}_colored.png")

            if os.path.exists(colored_mask_path):
                mask_image = cv2.imread(colored_mask_path)
                if mask_image is not None:
                    # Наложить цветную маску с полупрозрачностью
                    composite_mask = cv2.addWeighted(composite_mask, 1.0, mask_image, 0.8, 0)

        # Сохранить композитную маску
        timestamp = int(time.time())
        composite_mask_filename = f"composite_mask_{file_id}_{timestamp}.png"
        composite_mask_path = os.path.join("tmp", composite_mask_filename)

        pil_composite_mask = Image.fromarray(composite_mask)
        pil_composite_mask.save(composite_mask_path)

        return {
            "success": True,
            "composite_mask": composite_mask_path
        }

    except Exception as e:
        print(f"Error creating composite mask: {e}")
        return {"success": False, "error": str(e)}


@app.get("/api/list_tmp_dirs")
async def list_tmp_dirs():
    """
    Получение списка директорий в папке tmp для поиска масок полигонов.
    """
    try:
        tmp_path = "tmp"

        if not os.path.exists(tmp_path):
            print("❌ Папка tmp не существует")
            return []

        dirs = []
        for item in os.listdir(tmp_path):
            item_path = os.path.join(tmp_path, item)
            if os.path.isdir(item_path):
                dirs.append(item)

        print(f"📁 Все директории в tmp: {dirs}")

        # Возвращаем только директории с масками
        mask_dirs = [d for d in dirs if d.startswith('masks_')]
        print(f"🎯 Директории с масками: {mask_dirs}")

        return mask_dirs

    except Exception as e:
        print(f"Error listing tmp directories: {e}")
        return []


@app.get("/export/excel/all")
async def export_all_excel():
    """Экспорт всех результатов в Excel."""
    try:
        print(f"📊 Запрос на экспорт всех результатов")

        # Собираем все обработанные результаты
        if not processed_results:
            print(f"❌ Нет обработанных результатов")
            return {"error": "Нет обработанных результатов. Сначала обработайте изображения."}

        print(f"📋 Найдено {len(processed_results)} обработанных файлов")

        # Создаем данные для отчета всех изображений
        all_data = create_all_images_data()

        print(f"📊 Собрано данных: {len(all_data['parts'])} файлов с деталями, {len(all_data['damages'])} файлов с повреждениями")

        if not all_data['parts'] and not all_data['damages']:
            print(f"❌ Нет данных для экспорта")
            return {"error": "Нет данных для экспорта."}

        # Создаем файл
        filename = f"all_reports_{int(time.time())}.xlsx"
        file_path = create_all_excel_report(all_data, filename)

        # Определяем MIME type
        if file_path.endswith('.xlsx'):
            media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            media_type = 'text/plain'

        # Возвращаем файл
        return FileResponse(
            path=file_path,
            filename=os.path.basename(file_path),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={os.path.basename(file_path)}"}
        )

    except Exception as e:
        print(f"❌ Ошибка при экспорте: {e}")
        return {"error": f"Ошибка экспорта: {str(e)}"}


@app.get("/export/excel/{file_id}")
async def export_single_excel(file_id: str):
    """Экспорт данных одного изображения в Excel."""
    try:
        # Получаем реальные данные из обработанных результатов
        if file_id not in processed_results:
            return {"error": f"Результаты для file_id '{file_id}' не найдены. Сначала обработайте изображение."}

        data = processed_results[file_id]

        # Создаем файл
        filename = f"damage_report_{file_id}_{int(time.time())}.xlsx"
        file_path = create_excel_report(data, filename)

        # Определяем MIME type в зависимости от расширения
        if file_path.endswith('.xlsx'):
            media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            media_type = 'text/plain'

        # Возвращаем файл
        return FileResponse(
            path=file_path,
            filename=os.path.basename(file_path),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={os.path.basename(file_path)}"}
        )

    except Exception as e:
        print(f"❌ Ошибка при экспорте: {e}")
        return {"error": f"Ошибка экспорта: {str(e)}"}


def create_all_images_data():
    """Создает данные для отчета по всем обработанным изображениям."""
    all_parts = []
    all_damages = []

    print(f"🔍 Собираем данные для отчета из {len(processed_results)} файлов")

    for file_id, result in processed_results.items():
        filename = result.get('filename', f'file_{file_id}').replace('.jpg', '').replace('.png', '')
        mode = result.get('mode', 'unknown')

        # Собираем детали и повреждения из всех полигонов
        polygons = result.get('polygons', [])
        parts = [p.get('class', 'Неизвестно') for p in polygons if p.get('source') == 'parts']
        damages = [p.get('class', 'Неизвестно') for p in polygons if p.get('source') == 'damage']

        print(f"📄 Файл {filename} (режим: {mode}): {len(polygons)} полигонов, {len(parts)} деталей, {len(damages)} повреждений")

        # Добавляем детали для этого файла
        if parts:
            all_parts.append({
                'filename': filename,
                'parts': parts
            })
            print(f"✅ Добавлены детали: {parts}")

        # Добавляем повреждения для этого файла
        if damages:
            all_damages.append({
                'filename': filename,
                'damages': damages
            })
            print(f"✅ Добавлены повреждения: {damages}")

    print(f"📊 Итог: {len(all_parts)} файлов с деталями, {len(all_damages)} файлов с повреждениями")

    return {
        'parts': all_parts,
        'damages': all_damages
    }


def create_all_excel_report(all_data, filename="all_reports.xlsx"):
    """Создание Excel отчета для всех обработанных изображений."""

    # Проверяем, установлен ли openpyxl
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")

    try:
        # Создаем workbook
        wb = Workbook()

        # Удаляем лист по умолчанию
        wb.remove(wb.active)

        # Простые стили
        header_font = Font(bold=True, size=12)
        data_font = Font(size=10)

        # === ЛИСТ 1: ДЕТАЛИ ===
        parts_sheet = wb.create_sheet("ДЕТАЛИ")

        # Заголовки
        parts_sheet['A1'] = "Название файла"
        parts_sheet['B1'] = "Детали"
        parts_sheet['A1'].font = header_font
        parts_sheet['B1'].font = header_font

        # Записываем данные по всем файлам
        current_row = 2
        for i, file_data in enumerate(all_data['parts']):
            filename = file_data['filename']
            parts = file_data['parts']

            # Записываем название файла и детали
            for j, part in enumerate(parts):
                if j == 0:
                    # Первая деталь с названием файла
                    parts_sheet.cell(row=current_row, column=1).value = filename
                    parts_sheet.cell(row=current_row, column=1).font = data_font
                    parts_sheet.cell(row=current_row, column=2).value = part
                    parts_sheet.cell(row=current_row, column=2).font = data_font
                else:
                    # Остальные детали без названия файла
                    parts_sheet.cell(row=current_row, column=2).value = part
                    parts_sheet.cell(row=current_row, column=2).font = data_font
                current_row += 1

            # Добавляем пустую строку между файлами (кроме последнего)
            if i < len(all_data['parts']) - 1:
                current_row += 1

        # === ЛИСТ 2: ПОВРЕЖДЕНИЯ ===
        damage_sheet = wb.create_sheet("ПОВРЕЖДЕНИЯ")

        # Заголовки
        damage_sheet['A1'] = "Название файла"
        damage_sheet['B1'] = "Повреждения"
        damage_sheet['A1'].font = header_font
        damage_sheet['B1'].font = header_font

        # Записываем данные по всем файлам
        current_row = 2
        for i, file_data in enumerate(all_data['damages']):
            filename = file_data['filename']
            damages = file_data['damages']

            # Записываем название файла и повреждения
            for j, damage in enumerate(damages):
                if j == 0:
                    # Первое повреждение с названием файла
                    damage_sheet.cell(row=current_row, column=1).value = filename
                    damage_sheet.cell(row=current_row, column=1).font = data_font
                    damage_sheet.cell(row=current_row, column=2).value = damage
                    damage_sheet.cell(row=current_row, column=2).font = data_font
                else:
                    # Остальные повреждения без названия файла
                    damage_sheet.cell(row=current_row, column=2).value = damage
                    damage_sheet.cell(row=current_row, column=2).font = data_font
                current_row += 1

            # Добавляем пустую строку между файлами (кроме последнего)
            if i < len(all_data['damages']) - 1:
                current_row += 1

        # Автоподбор ширины колонок для обоих листов
        for sheet in [parts_sheet, damage_sheet]:
            for col_num in range(1, 3):  # Для колонок A-B (1-2)
                max_length = 0
                column_letter = chr(64 + col_num)  # A=65, B=66

                # Перебираем все строки в колонке
                for row_num in range(1, sheet.max_row + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    try:
                        # Проверяем, что ячейка имеет значение
                        if hasattr(cell, 'value') and cell.value:
                            cell_value = str(cell.value)
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                    except:
                        pass

                # Устанавливаем ширину колонки
                if max_length > 0:
                    adjusted_width = (max_length + 2) * 1.2
                    sheet.column_dimensions[column_letter].width = min(adjusted_width, 30)

        # Сохраняем файл в tmp директорию
        tmp_dir = "tmp"
        if not os.path.exists(tmp_dir):
            os.makedirs(tmp_dir)

        full_path = os.path.join(tmp_dir, filename)
        wb.save(full_path)

        print(f"✅ Excel файл для всех изображений сохранен: {full_path}")
        print(f"   📄 Файлов с деталями: {len(all_data['parts'])}")
        print(f"   🔧 Файлов с повреждениями: {len(all_data['damages'])}")

        return full_path

    except Exception as e:
        print(f"❌ Ошибка создания Excel файла для всех изображений: {e}")
        raise e


@app.get("/health")
async def health_check():
    """Проверка состояния сервиса."""
    from core.model_manager import _models_initialized, _damage_models_initialized
    return {
        "status": "healthy",
        "service": "DCD Vision API",
        "models": {
            "parts_initialized": _models_initialized,
            "damage_initialized": _damage_models_initialized
        }
    }


@app.get("/models/status")
async def models_status():
    """Проверка статуса загрузки моделей."""
    from core.model_manager import _models_initialized, _damage_models_initialized
    return {
        "parts_models_ready": _models_initialized,
        "damage_models_ready": _damage_models_initialized,
        "all_models_ready": _models_initialized and _damage_models_initialized
    }


@app.post("/cleanup/temp")
async def cleanup_temp():
    """Ручная очистка временных файлов."""
    try:
        cleanup_temp_files()
        return {
            "success": True,
            "message": "Временные файлы очищены"
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


@app.post("/cleanup/results")
async def cleanup_results():
    """Ручная очистка результатов обработки."""
    try:
        global processed_results
        processed_results.clear()
        return {
            "success": True,
            "message": f"Результаты обработки очищены. Удалено {len(processed_results)} записей"
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }




def create_excel_report(data, filename="damage_report.xlsx"):
    """Создание простого Excel отчета с двумя листами: ДЕТАЛИ и ПОВРЕЖДЕНИЯ."""

    # Проверяем, установлен ли openpyxl
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")

    try:
        # Создаем workbook
        wb = Workbook()

        # Удаляем лист по умолчанию
        wb.remove(wb.active)

        # Простые стили
        header_font = Font(bold=True, size=12)
        data_font = Font(size=10)

        # === ЛИСТ 1: ДЕТАЛИ ===
        parts_sheet = wb.create_sheet("ДЕТАЛИ")

        # Заголовки для деталей
        parts_sheet['A1'] = "Название файла"
        parts_sheet['B1'] = "Детали"
        parts_sheet['A1'].font = header_font
        parts_sheet['B1'].font = header_font

        # Получаем список деталей
        polygons = data.get('polygons', [])
        parts_list = [p.get('class', 'Неизвестно') for p in polygons if p.get('source') == 'parts']

        # Записываем данные
        filename_clean = data.get('filename', 'Неизвестно').replace('.jpg', '').replace('.png', '')
        parts_sheet['A2'] = filename_clean

        # Записываем детали в столбец B, начиная со строки 2
        for i, part in enumerate(parts_list):
            parts_sheet.cell(row=i+2, column=2).value = part
            parts_sheet.cell(row=i+2, column=2).font = data_font

        # === ЛИСТ 2: ПОВРЕЖДЕНИЯ ===
        damage_sheet = wb.create_sheet("ПОВРЕЖДЕНИЯ")

        # Заголовки для повреждений
        damage_sheet['A1'] = "Название файла"
        damage_sheet['B1'] = "Повреждения"
        damage_sheet['A1'].font = header_font
        damage_sheet['B1'].font = header_font

        # Получаем список повреждений
        damage_list = [p.get('class', 'Неизвестно') for p in polygons if p.get('source') == 'damage']

        # Записываем данные
        damage_sheet['A2'] = filename_clean

        # Записываем повреждения в столбец B, начиная со строки 2
        for i, damage in enumerate(damage_list):
            damage_sheet.cell(row=i+2, column=2).value = damage
            damage_sheet.cell(row=i+2, column=2).font = data_font

        # Автоподбор ширины колонок для обоих листов
        for sheet in [parts_sheet, damage_sheet]:
            for col_num in range(1, 3):  # Для колонок A-B (1-2)
                max_length = 0
                column_letter = chr(64 + col_num)  # A=65, B=66

                # Перебираем все строки в колонке
                for row_num in range(1, sheet.max_row + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    try:
                        # Проверяем, что ячейка имеет значение
                        if hasattr(cell, 'value') and cell.value:
                            cell_value = str(cell.value)
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                    except:
                        pass

                # Устанавливаем ширину колонки
                if max_length > 0:
                    adjusted_width = (max_length + 2) * 1.2
                    sheet.column_dimensions[column_letter].width = min(adjusted_width, 30)

        # Сохраняем файл в tmp директорию
        tmp_dir = "tmp"
        if not os.path.exists(tmp_dir):
            os.makedirs(tmp_dir)

        full_path = os.path.join(tmp_dir, filename)
        wb.save(full_path)

        print(f"✅ Excel файл сохранен: {full_path}")
        print(f"   📄 Детали: {len(parts_list)} объектов")
        print(f"   🔧 Повреждения: {len(damage_list)} объектов")

        return full_path

    except Exception as e:
        print(f"❌ Ошибка создания Excel файла: {e}")
        raise e


def create_text_report(data, filename="damage_report.xlsx"):
    """Создание текстового отчета вместо Excel."""
    try:
        tmp_dir = "tmp"
        if not os.path.exists(tmp_dir):
            os.makedirs(tmp_dir)

        txt_filename = filename.replace('.xlsx', '.txt')
        txt_path = os.path.join(tmp_dir, txt_filename)

        mode_text = "Детали + Повреждения" if data.get('mode') == 'full-union' else ("Повреждения" if data.get('mode') == 'damage' else "Детали")

        with open(txt_path, 'w', encoding='utf-8') as f:
            f.write("АНАЛИЗ АВТОМОБИЛЯ - DCD Vision\n")
            f.write("=" * 50 + "\n\n")
            f.write("Информация об изображении:\n")
            f.write(f"Имя файла: {data.get('filename', filename)}\n")
            f.write(f"Дата анализа: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n")
            f.write(f"Тип анализа: {mode_text}\n\n")

            polygons = data.get('polygons', [])
            parts_count = sum(1 for p in polygons if p.get('source') == 'parts')
            damage_count = sum(1 for p in polygons if p.get('source') == 'damage')

            f.write("НАЙДЕННЫЕ ДЕТАЛИ АВТОМОБИЛЯ:\n")
            f.write("-" * 40 + "\n")
            if parts_count == 0:
                f.write("Детали не найдены\n")
            else:
                for i, polygon in enumerate(polygons):
                    if polygon.get('source') == 'parts':
                        bbox = polygon.get('bbox', [0, 0, 0, 0])
                        confidence = polygon.get('confidence', 0.0)
                        f.write(f"{i+1:2d}. Деталь: {polygon.get('class', 'Неизвестно')}\n")
                        f.write(f"    Уверенность: {confidence:.1f}%\n")
                        f.write(f"    Координаты: X={int(bbox[0])}, Y={int(bbox[1])}\n")
                        f.write(f"    Размер: {int(bbox[2]-bbox[0])}x{int(bbox[3]-bbox[1])} пикселей\n")
                        f.write("-" * 40 + "\n")

            f.write("\nНАЙДЕННЫЕ ПОВРЕЖДЕНИЯ АВТОМОБИЛЯ:\n")
            f.write("-" * 40 + "\n")
            if damage_count == 0:
                f.write("Повреждения не найдены\n")
            else:
                for i, polygon in enumerate(polygons):
                    if polygon.get('source') == 'damage':
                        bbox = polygon.get('bbox', [0, 0, 0, 0])
                        confidence = polygon.get('confidence', 0.0)
                        f.write(f"{i+1:2d}. Повреждение: {polygon.get('class', 'Неизвестно')}\n")
                        f.write(f"    Уверенность: {confidence:.1f}%\n")
                        f.write(f"    Координаты: X={int(bbox[0])}, Y={int(bbox[1])}\n")
                        f.write(f"    Размер: {int(bbox[2]-bbox[0])}x{int(bbox[3]-bbox[1])} пикселей\n")
                        f.write("-" * 40 + "\n")

            f.write("\nСТАТИСТИКА АНАЛИЗА:\n")
            f.write("-" * 40 + "\n")
            f.write(f"Всего деталей найдено: {parts_count}\n")
            f.write(f"Всего повреждений найдено: {damage_count}\n")
            f.write(f"Общее количество объектов: {parts_count + damage_count}\n")
            f.write("\n" + "=" * 50 + "\n")
            f.write("Отчет создан: " + datetime.now().strftime('%d.%m.%Y %H:%M:%S') + "\n")

        print(f"📄 Создан текстовый отчет: {txt_path}")
        return txt_path

    except Exception as e:
        print(f"❌ Ошибка создания текстового отчета: {e}")
        return None






if __name__ == "__main__":
    print("📡 Сервер будет доступен по адресу: http://0.0.0.0:8001")
    print("🔄 Автоматическая перезагрузка включена - изменения в коде применятся автоматически")
    print("🛑 Для остановки нажмите Ctrl+C")

    # Используем subprocess для запуска uvicorn с reload через строку импорта
    import subprocess
    import sys

    try:
        # Запускаем uvicorn как subprocess с правильными параметрами для reload
        cmd = [
            sys.executable, "-m", "uvicorn",
            "main:app",  # Строка импорта вместо объекта
            "--host", "0.0.0.0",
            "--port", "8001",
            "--reload",  # Включаем reload
            "--reload-dir", ".",  # Мониторим текущую директорию
            "--log-level", "info"
        ]

        print(f"🚀 Запуск команды: {' '.join(cmd)}")
        subprocess.run(cmd, check=True)

    except KeyboardInterrupt:
        print("\n🛑 Сервер остановлен пользователем")
    except subprocess.CalledProcessError as e:
        print(f"❌ Ошибка запуска сервера: {e}")
    except Exception as e:
        print(f"❌ Непредвиденная ошибка: {e}")